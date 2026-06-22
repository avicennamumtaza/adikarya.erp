"""Centralised Excel import/export helpers for Adikarya ERP.

The workbook layout is shared across all datasets:

- Row 1: title
- Row 2: generated timestamp
- Row 3: column headers
- Row 4+: data

Pandas handles the tabular read/write path and openpyxl applies the workbook
styling so exports double as import templates.
"""

import io
import json
from datetime import date, datetime
from decimal import Decimal

import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.finance.models import FinancialAccount, FinancialTransaction, FinanceCategory
from apps.inventory.models import Branch, Product
from apps.partner.models import Contact
from apps.service.models import ServiceTicket
from apps.transaction.models import TransactionDetail, TransactionHeader

TITLE_ROW = 1
META_ROW = 2
HEADER_ROW = 3
DATA_ROW = 4

NAVY = "0F172A"
WHITE = "FFFFFF"
MUTED = "64748B"

TITLE_FONT = Font(name="Calibri", bold=True, color=WHITE, size=14)
TITLE_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
TITLE_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

META_FONT = Font(name="Calibri", italic=True, color=MUTED, size=9)
DATA_FONT = Font(name="Calibri", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

MONEY_FMT = '#,##0.00'
DATE_FMT = "DD/MM/YYYY"
DATETIME_FMT = "DD/MM/YYYY HH:MM"


def _is_missing(value):
    if value is None:
        return True
    if isinstance(value, float):
        return pd.isna(value)
    if isinstance(value, (list, dict, tuple, set)):
        return False
    try:
        return bool(pd.isna(value))
    except Exception:
        return False


def _safe_text(value, default=""):
    if _is_missing(value):
        return default
    text = str(value).strip()
    return text if text else default


def _safe_int(value, default=0):
    if _is_missing(value):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _safe_decimal(value, default=Decimal("0")):
    if _is_missing(value):
        return default
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    try:
        return Decimal(str(value).replace(",", ""))
    except (TypeError, ValueError, ArithmeticError):
        return default


def _safe_bool(value):
    if isinstance(value, bool):
        return value
    if _is_missing(value):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "ya", "y", "t"}


def _safe_date(value):
    if _is_missing(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _safe_text(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _safe_datetime(value):
    if _is_missing(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = _safe_text(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in {"%d/%m/%Y", "%Y-%m-%d"}:
                return datetime.combine(parsed.date(), datetime.min.time())
            return parsed
        except ValueError:
            continue
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def _safe_json_text(value):
    if _is_missing(value):
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _parse_json_list(value):
    if _is_missing(value):
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        return value
    text = str(value).strip()
    if not text:
        return []
    try:
        parsed = json.loads(text)
        if isinstance(parsed, (list, dict)):
            return parsed
    except (TypeError, ValueError, json.JSONDecodeError):
        pass
    return [part.strip() for part in text.split(";") if part.strip()]


def _money_to_float(value):
    return float(_safe_decimal(value))


def _normalize_count(value):
    return _safe_int(value, default=0)


def _series_value(row, *names):
    for name in names:
        if name in row.index:
            value = row[name]
            if _is_missing(value):
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return value
    return None


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is None:
                continue
            if isinstance(cell.value, datetime):
                text = cell.value.strftime("%d/%m/%Y %H:%M")
            elif isinstance(cell.value, date):
                text = cell.value.strftime("%d/%m/%Y")
            else:
                text = str(cell.value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _style_sheet(ws, title, headers, money_cols=None, date_cols=None, datetime_cols=None):
    money_cols = set(money_cols or [])
    date_cols = set(date_cols or [])
    datetime_cols = set(datetime_cols or [])

    col_count = len(headers)
    end_col = get_column_letter(col_count)

    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=col_count)
    title_cell = ws.cell(row=TITLE_ROW, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = TITLE_ALIGN
    ws.row_dimensions[TITLE_ROW].height = 30

    ws.merge_cells(start_row=META_ROW, start_column=1, end_row=META_ROW, end_column=col_count)
    meta_cell = ws.cell(
        row=META_ROW,
        column=1,
        value=f"Generated: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}",
    )
    meta_cell.font = META_FONT
    meta_cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    ws.freeze_panes = f"A{DATA_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{end_col}{max(ws.max_row, HEADER_ROW)}"

    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=ws.max_row, max_col=col_count):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row >= DATA_ROW:
                cell.font = DATA_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx in money_cols:
        for cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=DATA_ROW, max_row=ws.max_row):
            for cell in cells:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right", vertical="top")

    for col_idx in date_cols:
        for cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=DATA_ROW, max_row=ws.max_row):
            for cell in cells:
                cell.number_format = DATE_FMT

    for col_idx in datetime_cols:
        for cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=DATA_ROW, max_row=ws.max_row):
            for cell in cells:
                cell.number_format = DATETIME_FMT

    _auto_width(ws)


def _build_workbook_response(sheets, filename):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet in sheets:
            frame = pd.DataFrame(sheet["rows"], columns=sheet["columns"])
            frame.to_excel(writer, sheet_name=sheet["sheet_name"], index=False, startrow=2)

    buffer.seek(0)
    workbook = load_workbook(buffer)
    for sheet in sheets:
        ws = workbook[sheet["sheet_name"]]
        _style_sheet(
            ws,
            sheet["title"],
            sheet["columns"],
            money_cols=sheet.get("money_cols"),
            date_cols=sheet.get("date_cols"),
            datetime_cols=sheet.get("datetime_cols"),
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _sheet_names(file_obj):
    file_obj.seek(0)
    excel = pd.ExcelFile(file_obj, engine="openpyxl")
    names = list(excel.sheet_names)
    excel.close()
    file_obj.seek(0)
    return names


def _read_sheet(file_obj, sheet_name):
    file_obj.seek(0)
    frame = pd.read_excel(file_obj, sheet_name=sheet_name, header=2, engine="openpyxl")
    file_obj.seek(0)
    return frame


def _resolve_contact(name, preferred_type=None):
    if not name or name == "-":
        return None
    qs = Contact.objects.filter(name=name)
    if preferred_type:
        match = qs.filter(contact_type=preferred_type).first()
        if match:
            return match
    return qs.first()


def _resolve_branch(name):
    if not name or name == "-":
        return None
    return Branch.objects.filter(name=name).first()


def _resolve_product(sku=None, name=None):
    if sku:
        product = Product.objects.filter(sku=sku).first()
        if product:
            return product
    if name and name != "-":
        return Product.objects.filter(name=name).first()
    return None


def _resolve_account(name):
    if not name:
        return None
    return FinancialAccount.objects.filter(name=name).first()


def _resolve_category(name):
    if not name:
        return None
    return FinanceCategory.objects.filter(name=name).first()


def _resolve_invoice(invoice_number):
    if not invoice_number:
        return None
    return TransactionHeader.objects.filter(invoice_number=invoice_number).first()


def _normalize_result(created, updated):
    return created, updated


def _transaction_export_rows(qs, counterparty_label):
    summary_headers = [
        "Tanggal",
        "No. Invoice",
        counterparty_label,
        "Cabang",
        "Metode Bayar",
        "Total",
        "Terbayar",
        "Outstanding",
        "Jatuh Tempo",
        "Status",
    ]
    item_headers = [
        "No. Invoice",
        "SKU",
        "Nama Produk",
        "Qty",
        "Harga",
        "HPP",
        "Subtotal",
    ]

    summary_rows = []
    item_rows = []

    qs = qs.select_related("contact", "branch").prefetch_related("items__product")
    for header in qs:
        outstanding = _safe_decimal(header.total_amount) - _safe_decimal(header.amount_paid)
        if outstanding < 0:
            outstanding = Decimal("0")

        summary_rows.append(
            {
                "Tanggal": header.created_at,
                "No. Invoice": header.invoice_number,
                counterparty_label: header.contact.name if header.contact else "-",
                "Cabang": header.branch.name if header.branch else "-",
                "Metode Bayar": header.get_payment_method_display(),
                "Total": _money_to_float(header.total_amount),
                "Terbayar": _money_to_float(header.amount_paid),
                "Outstanding": _money_to_float(outstanding),
                "Jatuh Tempo": header.due_date,
                "Status": "Lunas" if outstanding <= 0 else "Belum Lunas",
            }
        )

        for item in header.items.all():
            subtotal = _safe_decimal(item.price_at_trx) * _safe_decimal(item.qty)
            item_rows.append(
                {
                    "No. Invoice": header.invoice_number,
                    "SKU": item.product.sku if item.product else "",
                    "Nama Produk": item.product.name if item.product else "",
                    "Qty": _normalize_count(item.qty),
                    "Harga": _money_to_float(item.price_at_trx),
                    "HPP": _money_to_float(item.cost_at_trx),
                    "Subtotal": _money_to_float(subtotal),
                }
            )

    return summary_headers, item_headers, summary_rows, item_rows


def export_products(qs):
    headers = [
        "SKU",
        "Nama Produk",
        "Tipe",
        "Kategori",
        "Brand",
        "Docs",
        "Harga Beli (Avg)",
        "Harga Jual",
        "Min Stok",
        "Stok Total",
        "Perlu Restock",
        "Catatan",
    ]
    rows = []
    for product in qs.prefetch_related("branch_stocks"):
        stock_total = sum(stock.quantity for stock in product.branch_stocks.all())
        rows.append(
            {
                "SKU": product.sku,
                "Nama Produk": product.name,
                "Tipe": product.get_product_type_display(),
                "Kategori": product.category or "",
                "Brand": product.brand or "",
                "Docs": product.docs or "",
                "Harga Beli (Avg)": _money_to_float(product.base_price),
                "Harga Jual": _money_to_float(product.selling_price),
                "Min Stok": _normalize_count(product.min_stock),
                "Stok Total": _normalize_count(stock_total),
                "Perlu Restock": "Ya" if stock_total <= product.min_stock else "Tidak",
                "Catatan": product.notes or "",
            }
        )
    return _build_workbook_response(
        [
            {
                "sheet_name": "Products",
                "title": "Data Produk - Adikarya ERP",
                "columns": headers,
                "rows": rows,
                "money_cols": {7, 8},
            }
        ],
        "products.xlsx",
    )


def export_partners(qs):
    headers = ["Nama", "Tipe", "WhatsApp", "Instagram", "Facebook", "Email", "Alamat", "Saldo", "Status Keuangan"]
    rows = []
    for partner in qs:
        rows.append(
            {
                "Nama": partner.name,
                "Tipe": partner.get_contact_type_display(),
                "WhatsApp": partner.whatsapp,
                "Instagram": partner.instagram or "",
                "Facebook": partner.facebook or "",
                "Email": partner.email or "",
                "Alamat": partner.address or "",
                "Saldo": _money_to_float(partner.current_balance),
                "Status Keuangan": partner.status_keuangan,
            }
        )
    return _build_workbook_response(
        [
            {
                "sheet_name": "Partners",
                "title": "Data Partner - Adikarya ERP",
                "columns": headers,
                "rows": rows,
                "money_cols": {8},
            }
        ],
        "partners.xlsx",
    )


def export_sales(qs):
    summary_headers, item_headers, summary_rows, item_rows = _transaction_export_rows(qs, "Customer")
    return _build_workbook_response(
        [
            {
                "sheet_name": "Sales",
                "title": "Data Penjualan - Adikarya ERP",
                "columns": summary_headers,
                "rows": summary_rows,
                "money_cols": {6, 7, 8},
                "date_cols": {1, 9},
            },
            {
                "sheet_name": "Sales Items",
                "title": "Item Penjualan - Adikarya ERP",
                "columns": item_headers,
                "rows": item_rows,
                "money_cols": {5, 6, 7},
            },
        ],
        "sales.xlsx",
    )


def export_purchases(qs):
    summary_headers, item_headers, summary_rows, item_rows = _transaction_export_rows(qs, "Supplier")
    return _build_workbook_response(
        [
            {
                "sheet_name": "Purchases",
                "title": "Data Pembelian - Adikarya ERP",
                "columns": summary_headers,
                "rows": summary_rows,
                "money_cols": {6, 7, 8},
                "date_cols": {1, 9},
            },
            {
                "sheet_name": "Purchases Items",
                "title": "Item Pembelian - Adikarya ERP",
                "columns": item_headers,
                "rows": item_rows,
                "money_cols": {5, 6, 7},
            },
        ],
        "purchases.xlsx",
    )


def export_finance(qs):
    headers = [
        "Tanggal",
        "Ref. Number",
        "Tipe",
        "Akun Sumber",
        "Akun Tujuan",
        "Kategori",
        "Invoice Ref",
        "Jumlah",
        "Biaya Admin",
        "Catatan",
        "Void",
    ]
    rows = []
    for transaction in qs.select_related("source_account", "destination_account", "category", "ref_invoice"):
        rows.append(
            {
                "Tanggal": transaction.date,
                "Ref. Number": transaction.reference_number or "",
                "Tipe": transaction.get_transaction_type_display(),
                "Akun Sumber": str(transaction.source_account) if transaction.source_account else "",
                "Akun Tujuan": str(transaction.destination_account) if transaction.destination_account else "",
                "Kategori": str(transaction.category) if transaction.category else "",
                "Invoice Ref": transaction.ref_invoice.invoice_number if transaction.ref_invoice else "",
                "Jumlah": _money_to_float(transaction.amount),
                "Biaya Admin": _money_to_float(transaction.fee),
                "Catatan": transaction.note or "",
                "Void": "Ya" if transaction.is_void else "Tidak",
            }
        )
    return _build_workbook_response(
        [
            {
                "sheet_name": "Finance",
                "title": "Data Keuangan - Adikarya ERP",
                "columns": headers,
                "rows": rows,
                "money_cols": {8, 9},
                "datetime_cols": {1},
            }
        ],
        "finance.xlsx",
    )


def export_service(qs):
    headers = [
        "No. Tiket",
        "Tanggal Masuk",
        "Customer",
        "Cabang",
        "Tipe Perangkat",
        "Merek",
        "Nama Perangkat",
        "Serial Number",
        "Warna",
        "Kelengkapan",
        "Catatan Kelengkapan",
        "Kondisi",
        "Keluhan",
        "Catatan Invoice",
        "Garansi (hari)",
        "Diskon",
        "Persetujuan Customer",
        "Status",
        "Invoice Ref",
    ]
    rows = []
    for ticket in qs.select_related("customer", "branch", "transaction"):
        rows.append(
            {
                "No. Tiket": ticket.ticket_number,
                "Tanggal Masuk": ticket.checkin_date,
                "Customer": ticket.customer.name if ticket.customer else "-",
                "Cabang": ticket.branch.name if ticket.branch else "-",
                "Tipe Perangkat": ticket.device_type,
                "Merek": ticket.device_brand,
                "Nama Perangkat": ticket.device_name,
                "Serial Number": ticket.serial_number,
                "Warna": ticket.device_color,
                "Kelengkapan": _safe_json_text(ticket.completeness),
                "Catatan Kelengkapan": ticket.completeness_notes,
                "Kondisi": _safe_json_text(ticket.condition),
                "Keluhan": ticket.complaint,
                "Catatan Invoice": ticket.invoice_notes,
                "Garansi (hari)": _normalize_count(ticket.warranty_days),
                "Diskon": _money_to_float(ticket.discount_amount),
                "Persetujuan Customer": "Ya" if ticket.customer_agreement else "Tidak",
                "Status": ticket.get_status_display(),
                "Invoice Ref": ticket.transaction.invoice_number if ticket.transaction else "",
            }
        )
    return _build_workbook_response(
        [
            {
                "sheet_name": "Service",
                "title": "Data Servis - Adikarya ERP",
                "columns": headers,
                "rows": rows,
                "money_cols": {16},
                "date_cols": {2},
            }
        ],
        "service_tickets.xlsx",
    )


def import_products(file_obj):
    frame = _read_sheet(file_obj, "Products")
    created = updated = 0
    errors = []

    product_type_map = {
        "Produk Fisik": "PRODUCT",
        "Jasa/Layanan": "SERVICE",
        "PRODUCT": "PRODUCT",
        "SERVICE": "SERVICE",
        "Goods": "PRODUCT",
        "Service": "SERVICE",
    }

    for index, row in frame.iterrows():
        row_number = index + DATA_ROW
        sku = _safe_text(_series_value(row, "SKU"))
        if not sku:
            continue
        try:
            defaults = {
                "name": _safe_text(_series_value(row, "Nama Produk")),
                "product_type": product_type_map.get(_safe_text(_series_value(row, "Tipe"), "PRODUCT"), "PRODUCT"),
                "category": _safe_text(_series_value(row, "Kategori")),
                "brand": _safe_text(_series_value(row, "Brand")),
                "docs": _safe_text(_series_value(row, "Docs")),
                "base_price": _safe_decimal(_series_value(row, "Harga Beli (Avg)") or 0),
                "selling_price": _safe_decimal(_series_value(row, "Harga Jual") or 0),
                "min_stock": _normalize_count(_series_value(row, "Min Stok") or 0),
                "notes": _safe_text(_series_value(row, "Catatan")),
            }
            _, is_created = Product.objects.update_or_create(sku=sku, defaults=defaults)
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as exc:
            errors.append(f"Baris {row_number}: {exc}")

    return _normalize_result(created, updated), errors


def import_partners(file_obj):
    frame = _read_sheet(file_obj, "Partners")
    created = updated = 0
    errors = []

    type_map = {
        "Customer": "CUSTOMER",
        "Supplier": "SUPPLIER",
        "Keduanya": "BOTH",
        "CUSTOMER": "CUSTOMER",
        "SUPPLIER": "SUPPLIER",
        "BOTH": "BOTH",
    }

    for index, row in frame.iterrows():
        row_number = index + DATA_ROW
        name = _safe_text(_series_value(row, "Nama"))
        if not name:
            continue
        try:
            defaults = {
                "contact_type": type_map.get(_safe_text(_series_value(row, "Tipe"), "CUSTOMER"), "CUSTOMER"),
                "whatsapp": _safe_text(_series_value(row, "WhatsApp")),
                "instagram": _safe_text(_series_value(row, "Instagram"), default=None),
                "facebook": _safe_text(_series_value(row, "Facebook"), default=None),
                "email": _safe_text(_series_value(row, "Email"), default=None),
                "address": _safe_text(_series_value(row, "Alamat")),
                "current_balance": _safe_decimal(_series_value(row, "Saldo") or 0),
            }
            _, is_created = Contact.objects.update_or_create(name=name, defaults=defaults)
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as exc:
            errors.append(f"Baris {row_number}: {exc}")

    return _normalize_result(created, updated), errors


def _import_transaction_sheet(file_obj, summary_sheet, item_sheet, trx_type, counterparty_label, preferred_contact_type):
    summary_frame = _read_sheet(file_obj, summary_sheet)
    sheet_names = set(_sheet_names(file_obj))
    item_frame = _read_sheet(file_obj, item_sheet) if item_sheet in sheet_names else None

    created = updated = 0
    errors = []

    method_map = {"Tunai": "CASH", "Tempo/Hutang": "CREDIT", "CASH": "CASH", "CREDIT": "CREDIT"}

    for index, row in summary_frame.iterrows():
        row_number = index + DATA_ROW
        invoice_number = _safe_text(_series_value(row, "No. Invoice", "Invoice Number"))
        if not invoice_number:
            continue
        try:
            contact = _resolve_contact(_safe_text(_series_value(row, counterparty_label)), preferred_type=preferred_contact_type)
            branch = _resolve_branch(_safe_text(_series_value(row, "Cabang")))
            if not contact or not branch:
                errors.append(f"Baris {row_number}: Contact atau Branch tidak ditemukan")
                continue

            defaults = {
                "contact": contact,
                "branch": branch,
                "trx_type": trx_type,
                "payment_method": method_map.get(_safe_text(_series_value(row, "Metode Bayar"), "CASH"), "CASH"),
                "total_amount": _safe_decimal(_series_value(row, "Total") or 0),
                "amount_paid": _safe_decimal(_series_value(row, "Terbayar") or 0),
                "due_date": _safe_date(_series_value(row, "Jatuh Tempo")),
                "is_finalized": True,
            }

            _, is_created = TransactionHeader.objects.update_or_create(invoice_number=invoice_number, defaults=defaults)
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as exc:
            errors.append(f"Baris {row_number}: {exc}")

    if item_frame is not None and not item_frame.empty:
        invoice_col = "No. Invoice" if "No. Invoice" in item_frame.columns else "Invoice Number"
        grouped = {}
        for index, row in item_frame.iterrows():
            row_number = index + DATA_ROW
            invoice_number = _safe_text(_series_value(row, invoice_col))
            if not invoice_number:
                continue
            grouped.setdefault(invoice_number, []).append((row_number, row))

        for invoice_number, rows in grouped.items():
            header = _resolve_invoice(invoice_number)
            if not header:
                errors.append(f"Invoice {invoice_number}: header tidak ditemukan untuk item")
                continue

            header.items.all().delete()
            item_total = Decimal("0")

            for row_number, row in rows:
                try:
                    product = _resolve_product(
                        sku=_safe_text(_series_value(row, "SKU")),
                        name=_safe_text(_series_value(row, "Nama Produk", "Product")),
                    )
                    if not product:
                        errors.append(f"Baris {row_number}: produk tidak ditemukan")
                        continue
                    qty = _normalize_count(_series_value(row, "Qty") or 0)
                    price = _safe_decimal(_series_value(row, "Harga") or 0)
                    cost = _safe_decimal(_series_value(row, "HPP") or 0)
                    TransactionDetail.objects.create(
                        header=header,
                        product=product,
                        qty=qty,
                        price_at_trx=price,
                        cost_at_trx=cost,
                    )
                    item_total += price * qty
                except Exception as exc:
                    errors.append(f"Baris {row_number}: {exc}")

            if item_total and _safe_decimal(header.total_amount) == Decimal("0"):
                header.total_amount = item_total
                header.save(update_fields=["total_amount"])

    return _normalize_result(created, updated), errors


def import_sales(file_obj):
    return _import_transaction_sheet(file_obj, "Sales", "Sales Items", "SALE", "Customer", "CUSTOMER")


def import_purchases(file_obj):
    return _import_transaction_sheet(file_obj, "Purchases", "Purchases Items", "PURCHASE", "Supplier", "SUPPLIER")


def import_finance(file_obj):
    frame = _read_sheet(file_obj, "Finance")
    created = updated = 0
    errors = []

    type_map = {
        "Masuk (Pendapatan/Modal/Piutang)": "IN",
        "Keluar (Beban/Hutang/Investasi)": "OUT",
        "Transfer Antar Akun": "TRANSFER",
        "IN": "IN",
        "OUT": "OUT",
        "TRANSFER": "TRANSFER",
    }

    for index, row in frame.iterrows():
        row_number = index + DATA_ROW
        try:
            ref_num = _safe_text(_series_value(row, "Ref. Number"), default=None)
            defaults = {
                "transaction_type": type_map.get(_safe_text(_series_value(row, "Tipe"), "IN"), "IN"),
                "source_account": _resolve_account(_safe_text(_series_value(row, "Akun Sumber"))),
                "destination_account": _resolve_account(_safe_text(_series_value(row, "Akun Tujuan"))),
                "category": _resolve_category(_safe_text(_series_value(row, "Kategori"))),
                "ref_invoice": _resolve_invoice(_safe_text(_series_value(row, "Invoice Ref"))),
                "amount": _safe_decimal(_series_value(row, "Jumlah") or 0),
                "fee": _safe_decimal(_series_value(row, "Biaya Admin") or 0),
                "date": _safe_datetime(_series_value(row, "Tanggal")) or timezone.now(),
                "note": _safe_text(_series_value(row, "Catatan")),
                "is_void": _safe_bool(_series_value(row, "Void")),
            }

            if ref_num:
                _, is_created = FinancialTransaction.objects.update_or_create(reference_number=ref_num, defaults=defaults)
                if is_created:
                    created += 1
                else:
                    updated += 1
            else:
                FinancialTransaction.objects.create(reference_number=None, **defaults)
                created += 1
        except Exception as exc:
            errors.append(f"Baris {row_number}: {exc}")

    return _normalize_result(created, updated), errors


def import_service(file_obj):
    frame = _read_sheet(file_obj, "Service")
    created = updated = 0
    errors = []

    status_map = {
        "Diterima": "RECEIVED",
        "Pengecekan": "DIAGNOSING",
        "Menunggu Part": "WAITING",
        "Perbaikan": "REPAIRING",
        "Siap Diambil": "DONE",
        "Diambil": "PICKED",
        "Batal": "CANCELLED",
        "RECEIVED": "RECEIVED",
        "DIAGNOSING": "DIAGNOSING",
        "WAITING": "WAITING",
        "REPAIRING": "REPAIRING",
        "DONE": "DONE",
        "PICKED": "PICKED",
        "CANCELLED": "CANCELLED",
    }

    for index, row in frame.iterrows():
        row_number = index + DATA_ROW
        ticket_number = _safe_text(_series_value(row, "No. Tiket"))
        if not ticket_number:
            continue
        try:
            customer = _resolve_contact(_safe_text(_series_value(row, "Customer")), preferred_type="CUSTOMER")
            branch = _resolve_branch(_safe_text(_series_value(row, "Cabang")))
            if not customer or not branch:
                errors.append(f"Baris {row_number}: Customer atau Branch tidak ditemukan")
                continue

            transaction = _resolve_invoice(_safe_text(_series_value(row, "Invoice Ref"), default=None))
            defaults = {
                "customer": customer,
                "branch": branch,
                "checkin_date": _safe_date(_series_value(row, "Tanggal Masuk")),
                "device_type": _safe_text(_series_value(row, "Tipe Perangkat")),
                "device_brand": _safe_text(_series_value(row, "Merek")),
                "device_name": _safe_text(_series_value(row, "Nama Perangkat")),
                "serial_number": _safe_text(_series_value(row, "Serial Number")),
                "device_color": _safe_text(_series_value(row, "Warna")),
                "completeness": _parse_json_list(_series_value(row, "Kelengkapan")),
                "completeness_notes": _safe_text(_series_value(row, "Catatan Kelengkapan")),
                "condition": _parse_json_list(_series_value(row, "Kondisi")),
                "complaint": _safe_text(_series_value(row, "Keluhan")),
                "invoice_notes": _safe_text(_series_value(row, "Catatan Invoice")),
                "warranty_days": _normalize_count(_series_value(row, "Garansi (hari)") or 0),
                "discount_amount": _safe_decimal(_series_value(row, "Diskon") or 0),
                "customer_agreement": _safe_bool(_series_value(row, "Persetujuan Customer")),
                "status": status_map.get(_safe_text(_series_value(row, "Status"), "RECEIVED"), "RECEIVED"),
                "transaction": transaction,
            }

            _, is_created = ServiceTicket.objects.update_or_create(ticket_number=ticket_number, defaults=defaults)
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as exc:
            errors.append(f"Baris {row_number}: {exc}")

    return _normalize_result(created, updated), errors